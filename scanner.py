import os
import time
import datetime
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Глобальный флаг для остановки процесса
stop_scanning_flag = False

def select_scan_dir():
    """Выбор папки для сканирования"""
    path = filedialog.askdirectory()
    if path:
        scan_entry.delete(0, tk.END)
        scan_entry.insert(0, path)

def select_save_dir():
    """Выбор папки для сохранения отчета"""
    path = filedialog.askdirectory()
    if path:
        save_entry.delete(0, tk.END)
        save_entry.insert(0, path)

def trigger_stop():
    """Кнопка СТОП меняет флаг прерывания"""
    global stop_scanning_flag
    stop_scanning_flag = True
    progress_label.config(text="Остановка... Подождите.")
    stop_btn.config(state=tk.DISABLED)

def start_scanning_thread():
    """Запуск сканирования в отдельном потоке, чтобы GUI не зависал"""
    global stop_scanning_flag
    stop_scanning_flag = False
    
    scan_path = scan_entry.get()
    save_path = save_entry.get()
    file_name = name_entry.get().strip()
    
    # Валидация входных данных
    if not scan_path or not os.path.exists(scan_path):
        messagebox.showerror("Ошибка", "Укажите существующую папку для сканирования!")
        return
    if not save_path or not os.path.exists(save_path):
        messagebox.showerror("Ошибка", "Укажите существующую папку для сохранения!")
        return
    if not file_name:
        messagebox.showerror("Ошибка", "Имя файла не может быть пустым!")
        return

    # Очистка имени файла от запрещенных символов
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        file_name = file_name.replace(char, '')

    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'

    # Переключение кнопок интерфейса
    start_btn.config(state=tk.DISABLED)
    stop_btn.config(state=tk.NORMAL)
    progress_bar['value'] = 0
    progress_label.config(text="Подсчет элементов...")
    
    # Запускаем обработку в фоне
    threading.Thread(target=process_scanning, args=(scan_path, save_path, file_name), daemon=True).start()

def format_duration(seconds):
    """Преобразование секунд в красивую строку времени"""
    seconds = round(seconds, 2)
    if seconds < 60:
        return f"{seconds} сек."
    else:
        minutes = int(seconds // 60)
        rem_seconds = round(seconds % 60, 2)
        return f"{minutes} мин. {rem_seconds} сек."

def process_scanning(scan_path, save_path, file_name):
    """Фоновый процесс сбора данных и генерации Excel"""
    global stop_scanning_flag
    
    # Формируем полный путь к будущему файлу заранее для проверки
    result_file = os.path.join(save_path, file_name)
    
    # Проверка, не открыт ли уже файл в Excel
    if os.path.exists(result_file):
        try:
            # Пытаемся открыть файл в монопольном режиме на запись
            f = open(result_file, 'r+')
            f.close()
        except IOError:
            # Если файл заблокирован системой (открыт в Excel)
            messagebox.showerror("Файл занят", "Данный файл уже открыт — закройте его или выберите другое имя")
            progress_label.config(text="Ошибка: Файл открыт в другой программе")
            return reset_gui_state()

    # Засекаем точное время старта
    start_time = time.time()
    
    try:
        # 1. Подсчет элементов
        total_items = 0
        for root_dir, dirs, files in os.walk(scan_path):
            if stop_scanning_flag: return reset_gui_state()
            total_items += len(dirs) + len(files)
        total_items += 1 
        if total_items == 0: total_items = 1

        # 2. Инициализация Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Дерево файлов"
        
        # Настройка первой строки (заголовков)
        header_font = Font(name="Arial", size=11, bold=True)
        ws.append(["Тип", "Имя элемента", "Расширение", "Дата изменения", "Размер (МБ)", "Полный путь"])
        
        ws.row_dimensions.height = 30 
        header_alignment = Alignment(horizontal="center", vertical="top") 
        
        # Применяем стили СТРОГО к ячейкам первой строки (от A1 до F1)
        for col_idx in range(1, 7):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_alignment


        # Стили для остальных строк данных
        custom_alignment = Alignment(horizontal="center", vertical="top")
        root_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        folder_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        file_font = Font(name="Arial", size=10, color="002060")
        default_font = Font(name="Arial", size=10)

        root_depth = scan_path.count(os.sep)
        processed_items = 0

        # 3. Обход директорий
        for root_dir, dirs, files in os.walk(scan_path):
            if stop_scanning_flag:
                progress_label.config(text="Процесс прерван пользователем")
                return reset_gui_state(interrupted=True)

            current_depth = root_dir.count(os.sep) - root_depth
            indent = "    " * current_depth
            
            # Получаем дату изменения папки
            try:
                mtime = os.path.getmtime(root_dir)
                date_val = datetime.datetime.fromtimestamp(mtime)
            except Exception:
                date_val = ""

            # Запись папки / корня
            current_row = ws.max_row + 1
            if root_dir != scan_path:
                folder_name = os.path.basename(root_dir)
                ws.append(["Папка", f"{indent}📁 {folder_name}", "Папка", date_val, "", root_dir])
                for cell in ws[current_row]:
                    cell.fill = folder_fill
                    cell.font = default_font
            else:
                ws.append(["Корень", f"📁 {os.path.basename(scan_path)}", "Папка", date_val, "", scan_path])
                for cell in ws[current_row]:
                    cell.fill = root_fill
                    cell.font = Font(name="Arial", size=10, bold=True)
            
            # Применяем выравнивание и формат даты к записанной строке папки/корня
            ws.cell(row=current_row, column=3).alignment = custom_alignment
            ws.cell(row=current_row, column=4).alignment = custom_alignment
            ws.cell(row=current_row, column=5).alignment = custom_alignment
            
            if date_val:
                ws.cell(row=current_row, column=4).number_format = 'yyyy-mm-dd hh:mm:ss'

            processed_items += 1
            update_progress(processed_items, total_items)

            # Запись файлов
            file_indent = "    " * (current_depth + 1)
            for file in files:
                if stop_scanning_flag:
                    progress_label.config(text="Процесс прерван пользователем")
                    return reset_gui_state(interrupted=True)

                file_path = os.path.join(root_dir, file)
                
                # Получаем расширение
                _, ext = os.path.splitext(file)
                ext = ext.lower() if ext else "Без расширения"
                
                # Получаем дату и размер файла
                try:
                    mtime = os.path.getmtime(file_path)
                    date_val = datetime.datetime.fromtimestamp(mtime)
                except Exception:
                    date_val = ""

                try:
                    size_bytes = os.path.getsize(file_path)
                    size_mb = size_bytes / (1024 * 1024)
                    if size_bytes > 0 and size_mb < 0.01:
                        size_val = "< 0.01"
                    else:
                        size_val = round(size_mb, 2)
                except Exception:
                    size_val = "-"

                current_row = ws.max_row + 1
                ws.append(["Файл", f"{file_indent}📄 {file}", ext, date_val, size_val, file_path])
                for cell in ws[current_row]:
                    cell.font = file_font
                
                # Применяем выравнивание для столбцов C, D, E файлов
                ws.cell(row=current_row, column=3).alignment = custom_alignment
                ws.cell(row=current_row, column=4).alignment = custom_alignment
                ws.cell(row=current_row, column=5).alignment = custom_alignment
                
                if date_val:
                    ws.cell(row=current_row, column=4).number_format = 'yyyy-mm-dd hh:mm:ss'
                
                if isinstance(size_val, (int, float)):
                    ws.cell(row=current_row, column=5).number_format = '#,##0.00'

                processed_items += 1
                update_progress(processed_items, total_items)

        # 4. Включение фильтров и автоподгон ширины колонок
        progress_label.config(text="Применение автофильтра и форматирование...")
        
        max_row = ws.max_row
        ws.auto_filter.ref = f"A1:F{max_row}"

        # Автоширина колонок
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Повторная финальная проверка на случай, если пользователь открыл файл во время сканирования
        wb.save(result_file)
        
                # Вычисляем разницу во времени
        end_time = time.time()
        elapsed_time_str = format_duration(end_time - start_time)
        
        progress_label.config(text=f"Готово! Время выполнения: {elapsed_time_str}")
        messagebox.showinfo("Успех", f"Сканирование успешно завершено!\nВремя работы: {elapsed_time_str}\nФайл сохранен: {result_file}")
        
    except PermissionError:
        messagebox.showerror("Ошибка доступа", "Не удалось сохранить файл. Данный файл уже открыт — закройте его или выберите другое имя")
        progress_label.config(text="Ошибка сохранения")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Произошла непредвиденная ошибка: {e}")
        progress_label.config(text="Ошибка при выполнении")
        
    finally:
        reset_gui_state()

def update_progress(processed, total):
    """Обновление прогресс-бара из фонового потока"""
    percent = int((processed / total) * 100)
    progress_bar['value'] = percent
    progress_label.config(text=f"Выполнено: {percent}% ({processed}/{total})")

def reset_gui_state(interrupted=False):
    """Возврат кнопок в дефолтное состояние"""
    start_btn.config(state=tk.NORMAL)
    stop_btn.config(state=tk.DISABLED)
    if interrupted:
        messagebox.showwarning("Прервано", "Сканирование было остановлено пользователем. Файл не сохранен.")

# --- Графический интерфейс (GUI) ---
root = tk.Tk()
root.title("Сканер папок в Excel (Многопоточный)")
root.geometry("550x420")
root.resizable(False, False)

# 1. Папка сканирования
tk.Label(root, text="Папка для сканирования:", font=("Arial", 10)).pack(anchor="w", padx=15, pady=(15, 2))
scan_frame = tk.Frame(root)
scan_frame.pack(fill="x", padx=15)
scan_entry = tk.Entry(scan_frame, font=("Arial", 10))
scan_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
tk.Button(scan_frame, text="Обзор...", command=select_scan_dir).pack(side="right")

# 2. Папка сохранения
tk.Label(root, text="Папка для сохранения файла:", font=("Arial", 10)).pack(anchor="w", padx=15, pady=(15, 2))
save_frame = tk.Frame(root)
save_frame.pack(fill="x", padx=15)
save_entry = tk.Entry(save_frame, font=("Arial", 10))
save_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
tk.Button(save_frame, text="Обзор...", command=select_save_dir).pack(side="right")

# 3. Имя файла Excel
tk.Label(root, text="Имя Excel файла (без расширения или с .xlsx):", font=("Arial", 10)).pack(anchor="w", padx=15, pady=(15, 2))
name_entry = tk.Entry(root, font=("Arial", 10))
name_entry.pack(fill="x", padx=15)
name_entry.insert(0, "file_tree")

# 4. Индикатор прогресса
tk.Label(root, text="Прогресс:", font=("Arial", 10)).pack(anchor="w", padx=15, pady=(15, 2))
progress_bar = ttk.Progressbar(root, orient="horizontal", mode="determinate")
progress_bar.pack(fill="x", padx=15)

progress_label = tk.Label(root, text="Ожидание запуска...", font=("Arial", 9, "italic"), fg="gray")
progress_label.pack(anchor="w", padx=15, pady=(2, 0))

# 5. Кнопки управления (СТАРТ и СТОП)
btn_frame = tk.Frame(root)
btn_frame.pack(fill="x", padx=15, pady=(20, 15))

start_btn = tk.Button(btn_frame, text="СТАРТ", font=("Arial", 11, "bold"), bg="#4CAF50", fg="white", height=2, width=22, command=start_scanning_thread)
start_btn.pack(side="left", fill="x", expand=True, padx=(0, 5))

stop_btn = tk.Button(btn_frame, text="СТОП", font=("Arial", 11, "bold"), bg="#F44336", fg="white", disabledforeground="white", height=2, width=22, state=tk.DISABLED, command=trigger_stop)
stop_btn.pack(side="right", fill="x", expand=True, padx=(5, 0))

root.mainloop()
